import os, json
from openpyxl import load_workbook
from os.path import join, isdir


class DataAnalysis():

    def __init__(self):
        self.errors = []
        if not isdir('S:/'):
            self.errors.append('S:/ Drive Not Connected')
        self.scan_project_directories()
        self.excel_database_analysis()
        self.load_analytics_settings()

    def scan_project_directories(self):
       
        if not self.errors:
            init_path = 'S:/PIERPONT MAINTENANCE & SERVICE/SERVICE CONTRACTS & PROPOSALS'
            # Maintenance Folders
            self.scanned_maintenance_jobs = {folder:folder for folder in os.listdir(init_path) if isdir(join(init_path, folder)) and folder[:3] != '000' and folder[3:6] == '-MS'}
            
            init_path = 'P:/'
            # Construction Folder
            scanned_construction_jobs = {}

            year_folder_list = [f for f in os.listdir(init_path) if isdir(join(init_path, f)) and f[:8] == 'Projects' and int(f[-4:]) >= 2021]

            for year in year_folder_list:

                for f in os.listdir(f'{init_path}/{year}'):
                    
                    if isdir(join(f'{init_path}/{year}', f)) and f[:3] != '000':

                        if f[3] == '-' or f[3] == '_':
                            name = f'{f[5:7]}-{f[:3]}_{f[8:]}'
                        else:
                            name = f'{f[6:8]}-{f[:4]}_{f[9:]}'
                        scanned_construction_jobs.update({name:f})

            self.scanned_construction_jobs = scanned_construction_jobs
        else:
            self.scanned_construction_jobs = {}
            self.scanned_maintenance_jobs = {}

    # ---------------------------------------------------------

    def excel_database_analysis(self):
        maintenance_database_clients = {}
        try:
            wb = load_workbook("S:/PIERPONT MAINTENANCE & SERVICE/SERVICE CONTRACTS & PROPOSALS/Primary_Maintenance_&_Service_Client_List.xlsx")  # Work Book
            ws = wb['maintenance-clients']  # Work Sheet
            job_numbers = ws['A']  
            names = ws['B']
            addresses = ws['C']
            frequency = ws['D']
            cost = ws['E']
            service_hourly = ws['F']

        except OSError as e:
            job_numbers = []
            self.errors.append('Maintenance Client Excel File Not Located in Proper Location')

        for x in range(1,len(job_numbers)):
            
            maintenance_database_clients.update({job_numbers[x].value:{
                'Name': names[x].value,
                'Address': addresses[x].value,
                'Frequency':frequency[x].value,
                'Cost_Per_Visit':cost[x].value,
                'Service_Rate':service_hourly[x].value}})
            

        self.maintenance_database_clients = maintenance_database_clients

    # ---------------------------------------------------------

    def load_analytics_settings(self):

        path = "S:/PIERPONT MAINTENANCE & SERVICE/Data_Storage/File_System_Scans/pierpont_analytics_settings.json"

        if not self.errors:
            try:
                with open(path, 'r') as j:
                    self.settings = json.load(j)
            except:
                self.settings = {}
                self.errors.append('Filter Settings File Not Located in Proper Location')
